from flask import Flask, request, send_file, jsonify, render_template
import pdfplumber
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import os

app = Flask(__name__)


def process_pdf_with_styles(pdf_file):
    """
    Конвертация таблицы из PDF в Excel.
    """
    with pdfplumber.open(pdf_file) as pdf:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PDF Table"

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        default_font = Font(name="Times New Roman", size=11)
        bold_font = Font(name="Times New Roman", size=11, bold=True)

        column_widths = []
        headers_seen = set()
        total_sum = 0
        current_row = 1
        
        header_text = pdf.pages[0].extract_text().split("\n")
        header_text = [line for line in header_text if line.strip()]
        
        for row_index, line in enumerate(header_text):
            if row_index > 7:
                continue
            if row_index >= 5:
                sheet.merge_cells(start_row=row_index - 4, start_column=1, end_row=row_index - 4, end_column=11)
                cell = sheet.cell(row=row_index - 4, column=1)
                cell.value = line
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1

        for page in pdf.pages:
            tables = page.extract_tables()

            if not tables:
                continue

            for table in tables:
                for row in table:
                    first_cell = row[0].strip() if row[0] else ""
                    if "Түсу күні/ Дата\nпоступления" in first_cell:
                        if first_cell in headers_seen:
                            continue
                        headers_seen.add(first_cell)

                    for col_index, cell in enumerate(row):
                        cell_value = str(cell).strip() if cell else ""
                        excel_cell = sheet.cell(row=current_row + 1, column=col_index + 1)
                        excel_cell.value = cell_value
                        excel_cell.font = default_font
                        excel_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        excel_cell.border = thin_border
                        
                        if col_index == 8:
                            try:
                                total_sum += float(cell_value)
                            except ValueError:
                                pass

                        cell_width = len(cell_value) + 2
                        if len(column_widths) < col_index + 1:
                            column_widths.append(cell_width)
                        else:
                            column_widths[col_index] = max(column_widths[col_index], cell_width)

                    current_row += 1

        for i, width in enumerate(column_widths, start=1):
            column_letter = get_column_letter(i)
            sheet.column_dimensions[column_letter].width = width
            
        last_row = sheet.max_row + 2
        sum_label_cell = sheet.cell(row=last_row, column=8)
        sum_label_cell.value = "Итого:"
        sum_label_cell.font = bold_font
        sum_label_cell.alignment = Alignment(horizontal="center", vertical="center")

        sum_value_cell = sheet.cell(row=last_row, column=9)
        sum_value_cell.value = total_sum
        sum_value_cell.font = bold_font
        sum_value_cell.alignment = Alignment(horizontal="center", vertical="center")
        sum_value_cell.border = thin_border

        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        return excel_file


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "Файл не выбран"}), 400

    pdf_file = request.files['file']

    try:
        excel_file = process_pdf_with_styles(pdf_file)
        return send_file(
            excel_file,
            as_attachment=True,
            download_name="output.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True)
